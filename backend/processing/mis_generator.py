import calendar
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from models import Visit

HEADER_FILL = PatternFill("solid", fgColor="D9C6A5")
SECTION_FILL = PatternFill("solid", fgColor="7B8B6F")
TOTAL_FILL = PatternFill("solid", fgColor="F0EBE0")
MARGIN_FILL = PatternFill("solid", fgColor="5C8A5C")
TITLE_FONT = Font(name="Inter", bold=True, size=13, color="2D2D2D")
HEADER_FONT = Font(name="Inter", bold=True, size=9, color="2D2D2D")
SECTION_FONT = Font(name="Inter", bold=True, size=10, color="FFFFFF")
CELL_FONT = Font(name="Inter", size=9)
TOTAL_FONT = Font(name="Inter", bold=True, size=9)


def generate_mis(db: Session, year: int, month: int) -> BytesIO:
    month_name = calendar.month_name[month]
    days_in_month = calendar.monthrange(year, month)[1]
    days = list(range(1, days_in_month + 1))

    visits = db.query(Visit).filter(Visit.year_num == year, Visit.month_num == month).all()
    workers = sorted({v.care_worker for v in visits if v.care_worker != "Not Assigned"})
    clients = sorted({v.service_user_no for v in visits})

    wh: dict = {}
    wp: dict = {}
    ch: dict = {}
    cb: dict = {}

    for v in visits:
        d = v.day_num
        if v.care_worker != "Not Assigned":
            wh[(v.care_worker, d)] = wh.get((v.care_worker, d), 0) + v.duration_hrs
            wp[(v.care_worker, d)] = wp.get((v.care_worker, d), 0) + v.care_worker_pay
        ch[(v.service_user_no, d)] = ch.get((v.service_user_no, d), 0) + v.duration_hrs
        cb[(v.service_user_no, d)] = cb.get((v.service_user_no, d), 0) + v.service_user_billing

    wb = Workbook()
    ws = wb.active
    ws.title = f"MIS {month_name}"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions[get_column_letter(34)].width = 22
    for d in days:
        ws.column_dimensions[get_column_letter(1 + d)].width = 6
        ws.column_dimensions[get_column_letter(34 + d)].width = 6

    ws.merge_cells("A1:AF1")
    title = ws["A1"]
    title.value = f"Daily worksheet — {month_name} {year}"
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    def write_section(start_row, label, entity_list, data, base_col, fmt):
        ws.merge_cells(start_row=start_row, start_column=base_col, end_row=start_row, end_column=base_col + len(days))
        cell = ws.cell(row=start_row, column=base_col, value=label)
        cell.font = SECTION_FONT; cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 18

        hdr_row = start_row + 1
        ws.cell(row=hdr_row, column=base_col, value="Name").font = HEADER_FONT
        ws.cell(row=hdr_row, column=base_col).fill = HEADER_FILL
        for i, d in enumerate(days):
            c = ws.cell(row=hdr_row, column=base_col + 1 + i, value=d)
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")

        for r_off, entity in enumerate(entity_list):
            row = hdr_row + 1 + r_off
            ws.cell(row=row, column=base_col, value=entity).font = CELL_FONT
            for i, d in enumerate(days):
                val = data.get((entity, d), "")
                c = ws.cell(row=row, column=base_col + 1 + i)
                if val:
                    c.value = round(val, 2); c.number_format = fmt
                c.font = CELL_FONT; c.alignment = Alignment(horizontal="center")

        total_row = hdr_row + 1 + len(entity_list)
        ws.cell(row=total_row, column=base_col, value="Total").font = TOTAL_FONT
        ws.cell(row=total_row, column=base_col).fill = TOTAL_FILL
        for i, d in enumerate(days):
            day_total = sum(data.get((e, d), 0) for e in entity_list)
            c = ws.cell(row=total_row, column=base_col + 1 + i)
            if day_total:
                c.value = round(day_total, 2); c.number_format = fmt
            c.font = TOTAL_FONT; c.fill = TOTAL_FILL
            c.alignment = Alignment(horizontal="center")
        return total_row + 2

    next_row = write_section(3, "Hours Worked", workers, wh, 1, "0.00")
    write_section(3, "Hours Paid (Amount £)", workers, wp, 34, "£#,##0.00")
    next_row = write_section(next_row, "Hours Billed", clients, ch, 1, "0.00")
    billed_start = next_row - (len(clients) + 4)
    write_section(billed_start, "Hours Billed (Amount £)", clients, cb, 34, "£#,##0.00")

    gm_row = next_row
    ws.cell(row=gm_row, column=34, value="Gross Margin").font = Font(name="Inter", bold=True, size=9, color="FFFFFF")
    ws.cell(row=gm_row, column=34).fill = MARGIN_FILL
    for i, d in enumerate(days):
        billing = sum(cb.get((e, d), 0) for e in clients)
        pay = sum(wp.get((e, d), 0) for e in workers)
        c = ws.cell(row=gm_row, column=35 + i)
        if billing or pay:
            c.value = round(billing - pay, 2); c.number_format = "£#,##0.00"
        c.font = Font(name="Inter", bold=True, size=9, color="FFFFFF")
        c.fill = MARGIN_FILL; c.alignment = Alignment(horizontal="center")

    summary_row = gm_row + 2
    total_revenue = sum(cb.values())
    total_cost = sum(wp.values())
    gross = total_revenue - total_cost
    summaries = [
        ("Total Revenue", total_revenue, "£#,##0.00"),
        ("Total Worker Cost", total_cost, "£#,##0.00"),
        ("Gross Margin (£)", gross, "£#,##0.00"),
        ("Gross Margin (%)", round(gross / total_revenue * 100, 1) if total_revenue else 0, "0.0%"),
        ("Total Care Hours", sum(ch.values()), "0.00"),
        ("Active Clients", len(clients), "0"),
        ("Active Care Workers", len(workers), "0"),
    ]
    ws.merge_cells(start_row=summary_row, start_column=34, end_row=summary_row, end_column=36)
    ws.cell(row=summary_row, column=34, value="Monthly Summary").font = SECTION_FONT
    ws.cell(row=summary_row, column=34).fill = SECTION_FILL
    for i, (label, value, fmt) in enumerate(summaries):
        r = summary_row + 1 + i
        ws.cell(row=r, column=34, value=label).font = TOTAL_FONT
        c = ws.cell(row=r, column=35, value=value)
        c.number_format = fmt; c.font = TOTAL_FONT
        c.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "B4"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf
